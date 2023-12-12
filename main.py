from flask import *
from auth import auth
from DAO import DAO
from datetime import date
from imagens_connection import insert_image
import openpyxl
import os

app = Flask(__name__)
app.register_blueprint(auth)

app.secret_key = b'_5#y2L"F4Q8z\n\xec]/'

def allowed_file(filename):
    UPLOAD_FOLDER = 'ocorrencias'
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'webp'}
    app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/home")
def home():
    if "loggedin" in session and session["loggedin"] is True:
        if session["pessoa_permissao"]:
            exportar_excel = '<li><a href="/exportar_excel">Exportar Excel</a></li>'
            return render_template("Templatedic.html", nome=session["nome"].split()[0], execel_export = exportar_excel)
        else:
            return render_template("Templatedic.html", nome=session["nome"].split()[0])
    else:
        flash("Logue para acessar esta página.")
        return redirect(url_for("auth.login"))


@app.route("/")
def index():
    return render_template("index.html")


@app.route("/depoimento")
def depoimento():
    if "loggedin" in session and session["loggedin"] is True:
        return render_template("Depoimento.html")
    else:
        flash("Logue para acessar esta página.")
        return redirect(url_for("auth.login"))


@app.route("/depoimento", methods=["POST"])
def depoimento_post():
    bd = DAO("ocorrencia")
    objDB = bd.ocorrencia()


    ocorrencia_descricao = request.form.get("depoimento")
    id_tipo_ocorrencia_fk = request.form.get("id_tipo_ocorrencia_fk")
    id_sub_lugar_fk = request.form.get("id_sub_lugar_fk")
    id_pessoa_fk = session["id"]
    ocorrencia_status = 0
    ocorrencia_data_registro = date.today()
    ocorrencia_data = "PLACEHOLDER"
    imagem = request.files['imagem']


    objDB.ocorrencia_descricao = ocorrencia_descricao
    objDB.id_tipo_ocorrencia_fk = id_tipo_ocorrencia_fk
    objDB.id_sub_lugar_fk = id_sub_lugar_fk
    objDB.id_pessoa_fk = id_pessoa_fk
    objDB.ocorrencia_status = ocorrencia_status
    objDB.ocorrencia_data_registro = ocorrencia_data_registro
    objDB.ocorrencia_data = ocorrencia_data
    bd.create(objDB)

    # captura a imagem da rota
    if imagem == None:
        return flash('message sem arquivo')

    if imagem and allowed_file(imagem.filename):
        novo_nome = f'ocorencia_num.' + imagem.filename.rsplit('.', 1)[1].lower()
        imagem.save(os.path.join(app.config['UPLOAD_FOLDER'], novo_nome))
        insert_image(os.getcwd() + rf'\ocorrencias\{novo_nome}', novo_nome)
        os.remove(os.getcwd() + rf'\ocorrencias\{novo_nome}')
    else:
        return flash('message use somente PNG, JPG ou WEBP')
    # ------------------------------------------------------------------------

    flash("Ocorrência registrada com sucesso!")



    return redirect(url_for("home"))

@app.route("/estatisticas")
def estatisticas():

    if session["pessoa_permissao"]:
            exportar_excel = '<li><a href="/exportar_excel">Exportar Excel</a></li>'
    
    Ocorrencias = DAO("ocorrencia")

    result = Ocorrencias.consultarEstatisticastipo()
    tipo_labels = [row.tipo_ocorrencia_nome for row in result]
    tipo_values = [row.total for row in result]

    lugar_result = Ocorrencias.consultarEstatisticaslugar()
    lugar_labels = [row.lugar_nome for row in lugar_result]
    lugar_values = [row.total for row in lugar_result]

    mes_result = Ocorrencias.consultarEstatisticasdata()
    mes_labels = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    mes_values = [row.total for row in mes_result]

    return render_template("graph_bar.html", tipo_labels=tipo_labels, lugar_labels=lugar_labels, mes_labels=mes_labels, tipo_values=tipo_values,lugar_values=lugar_values,mes_values=mes_values, nome=session['nome'].split()[0], execel_export = exportar_excel)

@app.route("/exportar_excel")
def exportar_excel():
    if session['pessoa_permissao']:
        wb = openpyxl.Workbook()
        ws = wb.active
        Excel = DAO("ocorrencia")
        result = Excel.consultarOcorrencias()
        id_ocorrencia = [row.ID_OCORRENCIA for row in result]
        setor_list = [row.SETOR for row in result]
        tipo_ocorrencia_list = [row.TIPO_OCORRENCIA for row in result]
        pessoa_nome_list = [row.NOME for row in result]
        ocorrencia_descricao_list = [row.DESCRICAO for row in result]
        ocorrencia_data_list = [row.DATA_OCORRIDO for row in result]
        ocorrencia_data_registro_list = [row.DATA_REGISTRO for row in result]
        ocorrencia_status_list = [row.OCORRENCIA_STATUS for row in result]
        status_legivel_list = ['Em Aberto' if status == 0 else 'Resolvido' for status in ocorrencia_status_list]
        lugar_nome_list = [row.LOCAL for row in result]
        sublugar_nome_list = [row.SUB_LOCAL for row in result]
        headers = ["ID_OCORRENCIA","SETOR", "TIPO_OCORRENCIA", "NOME", "DESCRICAO", "DATA_OCORRIDO", "DATA_REGISTRO", "OCORRENCIA_STATUS", "LOCAL", "SUB_LOCAL"]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        for row_num, data in enumerate(zip(id_ocorrencia,setor_list, tipo_ocorrencia_list, pessoa_nome_list, ocorrencia_descricao_list,
                            ocorrencia_data_list, ocorrencia_data_registro_list, status_legivel_list, lugar_nome_list, sublugar_nome_list), 2):
            for col_num, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_num, value=value)
        wb.save("Ocorrencias.xlsx")
        wb.close()

        return send_file("Ocorrencias.xlsx",as_attachment=True)
    else:
        return redirect(url_for('home'))

#rotas de erro
@app.errorhandler(404)
def page_not_found(error):
    return render_template('page_not_found.html')

@app.errorhandler(Exception)
def internal_server_error(error):
    return render_template('500_error.html')



if __name__ == "__main__":
    app.run(debug=True, use_reloader=True)
